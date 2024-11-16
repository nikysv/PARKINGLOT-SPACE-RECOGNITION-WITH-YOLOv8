import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel(file_path):
    # Conectar a la base de datos
    db_config = {
        'user': 'root',
        'password': 'mysql',
        'host': 'localhost',
        'database': 'parking_system'
    }
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    # Consultar los datos de la base de datos
    cursor.execute("SELECT id, space_number, llegada, salida, fecha, total FROM parking_log")
    rows = cursor.fetchall()

    # Crear un libro de trabajo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parking Log"

    # Agregar los encabezados
    headers = ["ID", "Nro. Parqueo", "Llegada", "Salida", "Día", "Total (Bs)"]
    ws.append(headers)

    # Estilizar los encabezados
    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajustar el ancho de las columnas
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar los datos de la base de datos
    for row in rows:
        ws.append(row)

    # Aplicar filtro en los encabezados
    ws.auto_filter.ref = ws.dimensions

    # Guardar el archivo Excel en la ruta seleccionada
    wb.save(file_path)
    print(f"Archivo Excel generado: {file_path}")

    # Cerrar conexión
    cursor.close()
    conn.close()
